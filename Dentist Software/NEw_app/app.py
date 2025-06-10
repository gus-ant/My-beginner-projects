from flask import Flask, render_template, request, redirect, url_for, session  # Import the session object
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
app.secret_key = 'WK22'

# Verifica se o arquivo Excel já existe
if os.path.exists("clinic.xlsx"):
    wb = load_workbook("clinic.xlsx")
else:
    wb = Workbook()

# Inicializa as variáveis de planilha
ws_patient = wb["Patients"] if "Patients" in wb.sheetnames else wb.create_sheet("Patients", 0)
ws_appointment = wb["Appointments"] if "Appointments" in wb.sheetnames else wb.create_sheet("Appointments", 1)
ws_medical_history = wb["Medical History"] if "Medical History" in wb.sheetnames else wb.create_sheet("Medical History", 2)
ws_treatment_plan = wb["Treatment Plans"] if "Treatment Plans" in wb.sheetnames else wb.create_sheet("Treatment Plans", 3)
ws_prescription = wb["Prescriptions"] if "Prescriptions" in wb.sheetnames else wb.create_sheet("Prescriptions", 4)
ws_material_used = wb["Materials Used"] if "Materials Used" in wb.sheetnames else wb.create_sheet("Materials Used", 5)
ws_purchase_request = wb["Purchase Requests"] if "Purchase Requests" in wb.sheetnames else wb.create_sheet("Purchase Requests", 6)

# Define the columns for each entity
patient_columns = ["ID", "Name", "Surname", "Birth Date", "Sex", "Address", "Phone", "Email"]
appointment_columns = ["ID", "Patient ID", "Date and Time", "Type", "Status", "Observations"]
medical_history_columns = ["ID", "Patient ID", "Date", "Type", "Description", "Observations"]
treatment_plan_columns = ["ID", "Patient ID", "Start Date", "End Date", "Description", "Observations"]
prescription_columns = ["ID", "Patient ID", "Date", "Medication", "Dosage", "Observations"]
material_used_columns = ["ID", "Name", "Quantity Used", "Date Used", "Observations"]
purchase_request_columns = ["ID", "Material ID", "Quantity Requested", "Date Requested", "Status"]


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # Lógica de autenticação
        # Se autenticado, armazene o estado na sessão
        session['logged_in'] = True
        return redirect(url_for("index"))  # Redireciona para a página inicial após login
    return render_template("login.html")

@app.route("/")
def index():
    if 'logged_in' not in session:  # Verifica se o usuário está autenticado
        return redirect(url_for("login"))  # Redireciona para a página de login se não estiver autenticado
    return render_template("index.html")

@app.route("/patients", methods=["GET", "POST"])
def patients():
    if request.method == "POST":
        patient_id = request.form["patient_id"]
        name = request.form["name"]
        surname = request.form["surname"]
        birth_date = request.form["birth_date"]
        sex = request.form["sex"]
        address = request.form["address"]
        phone = request.form["phone"]
        email = request.form["email"]
        patient_data = [patient_id, name, surname, birth_date, sex, address, phone, email]
        ws_patient.append(patient_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("patients"))
    patients = []
    for row in ws_patient.iter_rows(values_only=True):
        patients.append(row)
    return render_template("patients.html", patients=patients)

@app.route("/appointments", methods=["GET", "POST"])
def appointments():
    if request.method == "POST":
        patient_id = request.form["patient_id"]
        date_time = request.form["date_time"]
        type = request.form["type"]
        status = request.form["status"]
        observations = request.form["observations"]
        appointment_data = [None, patient_id, date_time, type, status, observations]
        ws_appointment.append(appointment_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("appointments"))
    appointments = []
    for row in ws_appointment.iter_rows(values_only=True):
        appointments.append(row)
    return render_template("appointments.html", appointments=appointments)

# Add more routes for medical history, treatment plans, prescriptions, materials used, and purchase requests


from flask import Flask, render_template, request, redirect, url_for, session
from werkzeug.utils import secure_filename
import os

# Add this new configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Add this new function to check allowed file extensions
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/medical_history", methods=["GET", "POST"])
def medical_history():
    if request.method == "POST":
        patient_id = request.form["patient_id"]
        date = request.form["date"]
        type = request.form["type"]
        description = request.form["description"]
        observations = request.form["observations"]
        
        # Handle file upload
        file = request.files['file_upload']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
        else:
            file_path = None

        medical_history_data = [None, patient_id, date, type, description, observations, file_path]
        ws_medical_history.append(medical_history_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("medical_history"))

    medical_histories = []
    for row in ws_medical_history.iter_rows(values_only=True):
        medical_histories.append(row)
    return render_template("medical_history.html", medical_histories=medical_histories)

@app.route("/treatment_plans", methods=["GET", "POST"])
def treatment_plans():
    if request.method == "POST":
        patient_id = request.form["patient_id"]
        start_date = request.form["start_date"]
        end_date = request.form["end_date"]
        description = request.form["description"]
        observations = request.form["observations"]
        treatment_plan_data = [None, patient_id, start_date, end_date, description, observations]
        ws_treatment_plan.append(treatment_plan_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("treatment_plans"))
    treatment_plans = []
    for row in ws_treatment_plan.iter_rows(values_only=True):
        treatment_plans.append(row)
    return render_template("treatment_plans.html", treatment_plans=treatment_plans)

@app.route("/prescriptions", methods=["GET", "POST"])
def prescriptions():
    if request.method == "POST":
        patient_id = request.form["patient_id"]
        date = request.form["date"]
        medication = request.form["medication"]
        dosage = request.form["dosage"]
        observations = request.form["observations"]
        prescription_data = [None, patient_id, date, medication, dosage, observations]
        ws_prescription.append(prescription_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("prescriptions"))
    prescriptions = []
    for row in ws_prescription.iter_rows(values_only=True):
        prescriptions.append(row)
    return render_template("prescriptions.html", prescriptions=prescriptions)

@app.route("/materials_used", methods=["GET", "POST"])
def materials_used():
    if request.method == "POST":
        name = request.form["name"]
        quantity_used = request.form["quantity_used"]
        date_used = request.form["date_used"]
        observations = request.form["observations"]
        material_used_data = [None, name, quantity_used, date_used, observations]
        ws_material_used.append(material_used_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("materials_used"))
    materials_used = []
    for row in ws_material_used.iter_rows(min_row=2, values_only=True):  # Começa a partir da segunda linha
        materials_used.append(row)  # Cada 'row' deve conter os dados na posição correta
    return render_template("materials_used.html", materials_used=materials_used)

@app.route("/purchase_requests", methods=["GET", "POST"])
def purchase_requests():
    if request.method == "POST":
        material_id = request.form["material_id"]
        quantity_requested = request.form["quantity_requested"]
        date_requested = request.form["date_requested"]
        status = request.form["status"]
        purchase_request_data = [None, material_id, quantity_requested, date_requested, status]
        ws_purchase_request.append(purchase_request_data)
        wb.save("clinic.xlsx")
        return redirect(url_for("purchase_requests"))

    purchase_requests = []
    for row in ws_purchase_request.iter_rows(min_row=2, values_only=True):  # Start from the second row
        purchase_requests.append(row)  # Each 'row' should contain the actual data

    print(purchase_requests)  # Debugging line to check the contents
    return render_template("purchase_requests.html", purchase_requests=purchase_requests)

@app.route("/update_purchase_request_status/<int:id>", methods=["POST"])
def update_purchase_request_status(id):
    new_status = request.form["new_status"]
    # Atualiza o status na planilha de solicitações de compra
    for row in ws_purchase_request.iter_rows(min_row=2, values_only=False):
        if row[0].value == id:  # Verifica se o ID corresponde
            row[4].value = new_status  # Atualiza o status
            break
    wb.save("clinic.xlsx")  # Salva as alterações no arquivo Excel
    return redirect(url_for("purchase_requests"))  # Redireciona para a página de solicitações de compra

@app.route("/update_material_status/<int:id>", methods=["POST"])
def update_material_status(id):
    new_status = request.form["new_status"]
    # Atualiza o status na planilha de materiais usados
    for row in ws_material_used.iter_rows(min_row=2, values_only=False):
        if row[0].value == id:  # Verifica se o ID corresponde
            row[4].value = new_status  # Atualiza o status
            break
    wb.save("clinic.xlsx")  # Salva as alterações no arquivo Excel
    return redirect(url_for("materials_used"))  # Redireciona para a página de materiais usados



@app.route("/dashboard")
def dashboard():
    # Coleta dados de cada planilha
    patients_data = [row for row in ws_patient.iter_rows(values_only=True)]
    appointments_data = [row for row in ws_appointment.iter_rows(values_only=True)]
    medical_history_data = [row for row in ws_medical_history.iter_rows(values_only=True)]
    treatment_plans_data = [row for row in ws_treatment_plan.iter_rows(values_only=True)]
    prescriptions_data = [row for row in ws_prescription.iter_rows(values_only=True)]
    materials_used_data = [row for row in ws_material_used.iter_rows(values_only=True)]
    purchase_requests_data = [row for row in ws_purchase_request.iter_rows(values_only=True)]

    return render_template("dashboard.html", 
                           patients=patients_data, 
                           appointments=appointments_data, 
                           medical_histories=medical_history_data, 
                           treatment_plans=treatment_plans_data, 
                           prescriptions=prescriptions_data, 
                           materials_used=materials_used_data, 
                           purchase_requests=purchase_requests_data)


@app.route("/weekly_expenses", methods=["GET"])
def weekly_expenses():
    # Aqui você deve implementar a lógica para calcular os gastos semanais
    # Exemplo de dados fictícios para a tabela
    weekly_expenses_data = [
        ["Semana 1", 150.00],
        ["Semana 2", 200.00],
        ["Semana 3", 175.00],
        ["Semana 4", 220.00],
    ]
    return render_template("weekly_expenses.html", weekly_expenses=weekly_expenses_data)

@app.route("/edit_patient/<string:patient_id>", methods=["GET", "POST"])
def edit_patient(patient_id):
    if request.method == "POST":
        # Get updated data from form
        name = request.form["name"]
        surname = request.form["surname"]
        birth_date = request.form["birth_date"]
        sex = request.form["sex"]
        address = request.form["address"]
        phone = request.form["phone"]
        email = request.form["email"]
        
        # Update patient data in worksheet
        for row in ws_patient.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == patient_id:
                row[1].value = name
                row[2].value = surname
                row[3].value = birth_date
                row[4].value = sex
                row[5].value = address
                row[6].value = phone
                row[7].value = email
                break
                
        wb.save("clinic.xlsx")
        return redirect(url_for("patients"))
    
    # Get current patient data for the form
    patient_data = None
    for row in ws_patient.iter_rows(values_only=True):
        if str(row[0]) == patient_id:
            patient_data = row
            break
            
    return render_template("edit_patient.html", patient=patient_data)

if __name__ == "__main__":
    app.run(debug=True)